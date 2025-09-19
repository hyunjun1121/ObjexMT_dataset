"""
Quick fix for similarity sheets - just truncate to 2817 rows
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("Quick fix: Truncating similarity sheets to 2817 rows...")

# Load the workbook
wb = load_workbook('OBJEX_dataset_labeling.xlsx')

# Sheets to fix
similarity_sheets = [
    'similarity_gpt-4.1',
    'similarity_claude-sonnet-4-2025',
    'similarity_Qwen3-235B-A22B-fp8-',
    'similarity_moonshotaiKimi-K2-In',
    'similarity_deepseek-aiDeepSeek-',
    'similarity_gemini-2.5-flash'
]

for sheet_name in similarity_sheets:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get current row count
        max_row = ws.max_row

        if max_row > 2818:  # 2817 data rows + 1 header
            print(f"  {sheet_name}: Removing rows {2819} to {max_row}")

            # Delete rows from 2819 onwards
            ws.delete_rows(2819, max_row - 2818)

# Save
wb.save('OBJEX_dataset_labeling_truncated.xlsx')
print("\nSaved as: OBJEX_dataset_labeling_truncated.xlsx")

# Verify
xl = pd.ExcelFile('OBJEX_dataset_labeling_truncated.xlsx')
print("\nVerification:")
for sheet in similarity_sheets:
    df = pd.read_excel(xl, sheet)
    print(f"  {sheet}: {len(df)} rows")